import pandas as pd
from openpyxl import load_workbook
import matplotlib
from matplotlib import pyplot as pt

excel_file = pd.ExcelFile('book.xlsx')
input = 99003768
#input = "Undrajavarapu Vijay"
# input = "undrajavarapu.vijay@ltts.com"

df = pd.DataFrame()

for i in excel_file.sheet_names:
    df1 = pd.read_excel(excel_file,i)
    # print(df1)

    # check for three inputs ,if they are found then print entire row
    df1.set_index('PS_Number', inplace=True)

    result = df1.loc[input]
    print(result)
 #   list_call = []
 #   list_call.append(result)
 #   print(list)

df = pd.DataFrame({'PS_Number': ['A', 'B', 'C'], 'Team_No_1': [10, 30, 20]})
ax = df.plot.bar(x='PS_Number', y='Team_No_1', rot=0)
# pt.show()

# path = "book.xlsx"
# book = load_workbook(path)
#
# writer = pd.ExcelWriter(path, engine='openpyxl')
# writer.book = book
# if 'Mastersheet' in book.sheetnames:
#     ref = book['Mastersheet']
#     book.remove(ref)
#
# result.to_excel(writer, sheet_name='Mastersheet')
# #ax.to_excel(writer, sheet_name='Mastersheet')
#
# writer.save()
# writer.close()











# input = input("Enter Area: ")
# df_mini_first_sheet = pd.read_excel(excel_file, sheet_name=0)
# df_mini_second_sheet = pd.read_excel(excel_file, sheet_name=1)
# df_mini_third_sheet = pd.read_excel(excel_file, sheet_name=2)
# df_mini_fourth_sheet = pd.read_excel(excel_file, sheet_name=3)
# df_mini_fifth_sheet = pd.read_excel(excel_file, sheet_name=4)
# file3 = df_mini_first_sheet.merge(df_mini_second_sheet, on="Area_Name", how="left")
# file4 = file3.merge(df_mini_third_sheet, on="Area_Name", how="left")
# file5 = file4.merge(df_mini_fourth_sheet, on="Area_Name", how="left")
# file6 = file5.merge(df_mini_fifth_sheet, on="Area_Name", how="left")
#
# df_input = pd.DataFrame(file6, columns=['Area_Name', 'Year(2005)','Year(2004)'])
# df_input.set_index('Area_Name', inplace=True)
#
# result = df_input.loc[input]
# print(result)
#
# path = "datasheet_crime_excel.xlsx"
# book = load_workbook(path)
#
# writer = pd.ExcelWriter(path, engine='openpyxl')
# writer.book = book
# if 'Mastersheet' in book.sheetnames:
#     ref = book['Mastersheet']
#     book.remove(ref)
#
# result.to_excel(writer, sheet_name='Mastersheet')
#
# writer.save()
# writer.close()

