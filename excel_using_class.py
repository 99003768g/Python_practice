import pandas as pd
from openpyxl import load_workbook

excel_file = pd.ExcelFile('datasheet_crime_excel.xlsx')
sheets = excel_file.sheet_names

# input = input("Enter Area: ")
input = 'Goa'

def read_excel(sheets):
    df_mini_sheet = pd.read_excel(excel_file, sheet_name=sheets)
    return df_mini_sheet


def merge(self, df_mini_sheet, Area_Name):
    self.df_mini_sheet = df_mini_sheet
    self.Area_Name = Area_Name
    file6 = df_mini_sheet.merge(df_mini_sheet, on=Area_Name)
    return file6


df_input = pd.DataFrame(merge.file6, columns=['Area_Name', 'Year(2005)', 'Year(2004)'])
df_input.set_index('Area_Name', inplace=True)

result = df_input.loc[input]
print(result)

path = "datasheet_crime_excel.xlsx"
book = load_workbook(path)

writer = pd.ExcelWriter(path, engine='openpyxl')
writer.book = book
if 'Mastersheet' in book.sheetnames:
    ref = book['Mastersheet']
    book.remove(ref)

result.to_excel(writer, sheet_name='Mastersheet')

writer.save()
writer.close()
