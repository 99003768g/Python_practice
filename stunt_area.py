import openpyxl

path = "book.xlsx"

# To open the workbook , workbook object wb_obj is created
wb_obj = openpyxl.load_workbook(path)
sheets = wb_obj.sheetnames

#a= input('enter to search')
a = 99003768
for i in sheets:
    sb = wb_obj[i]

    for i in range(1, sb.max_row):
        for j in range(i, sb.max_column):
            if a == sb[23]:
                print("found")


    # for col in sb.iter_cols():
    #     for cell in col:
    #         if cell.value == 99003768:
    #             print(cell.value,cell)

 #   search for the keyword here

    # for row in sb.iter_rows():
    #      for cell in row:
    #          if cell.value == "Undrajavarapu Vijay":
    #             print("Found at cell", cell)

    # for col in sb.iter_cols():
    #     for cell in col:
    #         if cell.value == "Name":
    #             print("Found at cell", cell)
    #
    #
    #
    # for i in range(1, sb.max_row):
    #     if sb.cell(row=i,column=i).value == 'PS_Number':
    #         for j in range(i, sb.max_column):
    #             print(sb.cell(row=i, column=j).value)

































# import openpyxl
# from collections import defaultdict
# path = "book.xlsx"
# wb_obj = openpyxl.load_workbook(path)
# sheets = wb_obj.sheetnames
#
# Variable_Model_Path = defaultdict(list)
# Sheet_Name = wb_obj[sheets]
# keys = False
#
# # for row in Sheet_Name.iter_rows():
# #     if not keys:
# #         keys = [cell.value for cell in row]
# #         continue
# #     Row=[cell.value for cell in row]
# #     for cell_ind,cell in enumerate(Row):
# #         Variable_Model_Path[keys[cell_ind]].append( cell)
# # print(keys)