# open_workbook.py
# from openpyxl import load_workbook
# def open_workbook(path):
#     workbook = load_workbook(filename=path)
#     print(f'Worksheet names: {workbook.sheetnames}')
#     sheet = workbook.active
#     print(sheet)
#     print(f'The title of the Worksheet is: {sheet.title}')
#
#
# if __name__ == '__main__':
#     open_workbook('books.xlsx')

import openpyxl

path = "book.xlsx"

# To open the workbook , workbook object wb_obj is created
wb_obj = openpyxl.load_workbook(path)
sheets = wb_obj.sheetnames
# a= input('enter to search')
for i in sheets:
    sb = wb_obj[i]

# search for the keyword here

# for row in sb.iter_rows():
# for cell in row:
# f cell.value == "Undrajavarapu Vijay":
# print("Found at cell", cell)

    for inside in range(1, sb.max_row):
        for cell in sb.iter_rows():
            if cell.value == 99003768:
                print(sb.cell(row=i, column=cell).value)
# for i in range(1, sb.max_row):
# if sb.cell(row=i,column=i).value == 'Name':
# for j in range(i, sb.max_column):
# print(sb.cell(row=i, column=j).value)
